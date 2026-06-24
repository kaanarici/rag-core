"""Workbook loading helpers for the XLSX converter."""

from __future__ import annotations

import io
import logging
import zipfile
from dataclasses import dataclass, field
from typing import Any


@dataclass(frozen=True)
class XlsxWorkbooks:
    data: Any | None
    formula: Any | None = None
    chart: Any | None = None
    chart_map: dict[str, list[Any]] = field(default_factory=dict)
    error: str = ""


@dataclass(frozen=True)
class XlsxWorkbookHints:
    has_formula: bool
    has_chart: bool


def load_xlsx_workbooks(
    file_bytes: bytes,
    *,
    format_name: str,
    logger: logging.Logger,
) -> XlsxWorkbooks:
    from openpyxl import load_workbook

    try:
        data = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        logger.warning(
            "Failed to open Office document: format=%s error_type=%s",
            format_name,
            type(exc).__name__,
        )
        return XlsxWorkbooks(
            data=None,
            error="XLSX parse failed (%s)" % type(exc).__name__,
        )

    hints = _inspect_xlsx_package(file_bytes)
    formula = None
    if hints.has_formula:
        try:
            formula = load_workbook(
                io.BytesIO(file_bytes),
                data_only=False,
                read_only=True,
            )
        except Exception as exc:
            logger.debug(
                "Could not load formula workbook: format=%s error_type=%s",
                format_name,
                type(exc).__name__,
            )

    chart = None
    chart_map: dict[str, list[Any]] = {}
    if hints.has_chart:
        try:
            chart = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
            for chart_sheet in chart.worksheets:
                chart_map[chart_sheet.title] = list(getattr(chart_sheet, "_charts", []))
        except Exception as exc:
            logger.debug(
                "Could not load chart metadata: format=%s error_type=%s",
                format_name,
                type(exc).__name__,
            )

    return XlsxWorkbooks(
        data=data,
        formula=formula,
        chart=chart,
        chart_map=chart_map,
    )


def _inspect_xlsx_package(file_bytes: bytes) -> XlsxWorkbookHints:
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            names = archive.namelist()
            return XlsxWorkbookHints(
                has_formula=_has_formula_parts(archive, names),
                has_chart=any(
                    name.startswith("xl/charts/") and name.endswith(".xml")
                    for name in names
                ),
            )
    except Exception:
        return XlsxWorkbookHints(has_formula=True, has_chart=True)


def _has_formula_parts(archive: zipfile.ZipFile, names: list[str]) -> bool:
    for name in names:
        if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
            continue
        tail = b""
        with archive.open(name) as worksheet:
            while chunk := worksheet.read(65536):
                haystack = tail + chunk
                if (
                    b"<f>" in haystack
                    or b"<f " in haystack
                    or b"<f\t" in haystack
                    or b"<f\r" in haystack
                    or b"<f\n" in haystack
                ):
                    return True
                tail = haystack[-3:]
    return False


def close_xlsx_workbooks(workbooks: XlsxWorkbooks) -> None:
    for workbook in (workbooks.formula, workbooks.chart, workbooks.data):
        if workbook is None:
            continue
        try:
            workbook.close()
        except Exception:
            pass
