from __future__ import annotations

import asyncio
from io import BytesIO
from typing import Any

import pytest

from rag_core.documents.converters.xlsx_converter import XlsxConverter


def _workbook_bytes(*, formula: bool = False) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Signals"
    sheet.append(["Signal", "Value"])
    sheet.append(["retrieval_quality", 98])
    sheet.append(["latency", 42])
    if formula:
        sheet.append(["total", "=SUM(B2:B3)"])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_xlsx_simple_workbook_uses_only_read_only_data_load(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from openpyxl import load_workbook

    calls: list[tuple[bool, bool]] = []

    def recording_load_workbook(
        stream: Any,
        *,
        data_only: bool,
        read_only: bool,
    ) -> Any:
        calls.append((data_only, read_only))
        return load_workbook(stream, data_only=data_only, read_only=read_only)

    monkeypatch.setattr("openpyxl.load_workbook", recording_load_workbook)

    result = asyncio.run(
        XlsxConverter().convert(
            _workbook_bytes(),
            filename="signals.xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )

    assert "| retrieval_quality | 98 |" in result.content
    assert calls == [(True, True)]


def test_xlsx_formula_workbook_keeps_formula_load_lazy_without_chart_load(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from openpyxl import load_workbook

    calls: list[tuple[bool, bool]] = []

    def recording_load_workbook(
        stream: Any,
        *,
        data_only: bool,
        read_only: bool,
    ) -> Any:
        calls.append((data_only, read_only))
        return load_workbook(stream, data_only=data_only, read_only=read_only)

    monkeypatch.setattr("openpyxl.load_workbook", recording_load_workbook)

    result = asyncio.run(
        XlsxConverter().convert(
            _workbook_bytes(formula=True),
            filename="signals.xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )

    assert "=SUM(B2:B3)" in result.content
    assert calls == [(True, True), (False, True)]
